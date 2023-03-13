import Main as M

from selenium import webdriver

from selenium.webdriver.common.by import By

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

import pyautogui
import clipboard
import time
import os
import random
import string
import openpyxl as op


IDPW = "12121212"
WALLET_ADDRESS = ""

#==================이더메일 기록을 위한 기존 GMAIL정보==============
PATH_EXCEL = "C:/TEST/SHARDEUM/"
FILE_EXCLEL = "PC0_SHARDEUM.xlsx"

SHARDEUM_WB = op.load_workbook(PATH_EXCEL + FILE_EXCLEL)
SHARDEUM_WS = SHARDEUM_WB["SHARDEUM"]

print(SHARDEUM_WB.sheetnames)

#======================랜덤 샤디움 네이밍 생성======================

def SHARDEUM_NAME():
    SHARDEUM_NAME=8	# 문자의 개수(문자열의 크기)
    RANDOM_str = ""	# 문자열
    for i in range(SHARDEUM_NAME):
        RANDOM_str += str(random.choice(string.ascii_letters + string.digits))
    return(RANDOM_str)

##########################################################
#####################크롬 프로필 리스트화###################
##########################################################

# 크롬 계정데이터 모여있는 폴더 chrome://version
directory_chrome = r"C:\Users\thqud\AppData\Local\Google\Chrome\User Data"

# 해당 문자가 포함된 폴더만 리스트화 시키겠다 (프로필 뒤에 공백 필수)
include_text = 'Profile '

# 디렉토리명 필터링
DIR_PROFILE_LIST = (
    [d for d in os.listdir(directory_chrome) 
    if os.path.isdir(os.path.join(directory_chrome, d)) and include_text in d]
)
"""
#폴더명 리스트화
print(len(DIR_PROFILE_LIST))
for i in range(len(DIR_PROFILE_LIST)):
    print(DIR_PROFILE_LIST[i])
"""

##########################################################
#####################셀레니움 기본 세팅#####################
##########################################################
options = Options()
user_data = r"C:\Users\thqud\AppData\Local\Google\Chrome\User Data"
options.add_argument(f"user-data-dir={user_data}")

#=======================실행 / 반복 부분 ===========================
#몇번부터 실행?
ALL_RANGE_CHROME = len(DIR_PROFILE_LIST)

n = int(input("몇번부터 시작? 1 ~ %d: " %ALL_RANGE_CHROME))
print(type(n),n)
n -= 1 #사용자 편의를 위해 0부터 시작이 아닌 1부터 시작

##########################################################
#####################자동화 코드 시작######################
##########################################################

for i in range(len(DIR_PROFILE_LIST)):
    if(n < len(DIR_PROFILE_LIST)):    
        options.add_argument("--profile-directory=" + DIR_PROFILE_LIST[n])
        options.add_experimental_option("detach", True)  # 화면이 꺼지지 않고 유지
        options.add_argument("--start-maximized")  # 최대 크기로 시작

        service = Service(ChromeDriverManager().install()) # 웹드라이버 설치
        driver = webdriver.Chrome(service=service, options=options) # 웹드라이버 불러오기
        action = ActionChains(driver)
    else:
        print("샤디움 자동화 프로그램 종료")
        break #계정이 끝에 도달했을 때 프로그램 종료
    
    #try except문에서 효과적으로 나오기 위해 사용
    for j in range(1):
        # 0 - 메타마스크 미리 로그인 (팝업방지용도)
        try:
            driver.get("chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html") #웹사이트 이동
            M.TIME(4)
            action.send_keys(IDPW).key_down(Keys.ENTER).pause(3).perform()
            driver.find_element(By.CLASS_NAME,"selected-account__clickable").click()
            WALLET_ADDRESS = clipboard.paste() #지갑주소 변수에 저장
            action.reset_actions() #액션 종료
        except:
            print("# 0 Failed")
            break

        # 1 - 샤디움 2.x 연결
        try:
            driver.get("https://chainlist.org/?search=Shardeum") #웹사이트 이동
            M.TIME(3)
            (
            action.send_keys((Keys.TAB)*11).key_down(Keys.ENTER).key_up(Keys.ENTER).pause(7)
            .perform()
            )

            print(driver.window_handles)
            driver.switch_to.window(driver.window_handles[1]) #크롬 팝업창으로 이동
            M.TIME(5)
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary").click() #다음
            M.TIME(2)
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary.page-container__footer-button").click()#연결
            M.TIME(4)
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary").click()
            M.TIME(2)
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary").click()
            M.TIME(4)
        except:
            print("# 1 Failed")
            break

        # 2 - 샤디움 1.x 연결
        try:
            driver.switch_to.window(driver.window_handles[0]) #원래 크롬창으로 이동
            driver.get("https://chainlist.org/?search=Shardeum") #웹사이트 이동
            M.TIME(3)
            (
            action.send_keys((Keys.TAB)*8).key_down(Keys.ENTER).key_up(Keys.ENTER).pause(5)
            .perform()
            )
        except:
            print("# 2 Failed")
            break
            
        # 3 - 네트워크 추가 허용
        try:
            driver.get("chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html#confirmation")#네트워크 추가 허용 승인
            M.TIME(3)
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary").click()
            M.TIME(2)
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary").click() #네트워크전환
            M.TIME(4)
        except:
            print("# 3 Failed")
            break
            
        # 4 - 디스코드 초대수락
        try:
            driver.get("https://discord.com/invite/shardeum")#초대페이지 이동
            M.TIME(3)
            driver.find_element(By.CLASS_NAME,"contents-3ca1mk").click() #초대승인
            M.TIME(4)
            driver.get("https://discord.com/channels/933959587462254612/938398242788487168")#디스코드 채팅방 이동
            M.TIME(5)
            #driver.find_element(By.CLASS_NAME,"button-2Z42pt.button-f2h6uQ.lookFilled-yCfaCM.colorBrand-I6CyqQ.sizeMedium-2bFIHr.grow-2sR_-F").click() #Capito #아마 첫 로그인떄만 뜨는 창일겁니다
            M.TIME(2)
        except:
            print("# 4 Failed")
            break

        # 5 - 디스코드 미션 수행 1
        try:
            driver.find_element(By.CLASS_NAME,"channelMention.wrapper-1ZcZW-.iconMention-3WxjBe.iconMentionText-1_WCtN.mention.interactive").click() #Verrify
            M.TIME(3)
            driver.find_element(By.CLASS_NAME,"reactionCount-26U4As").click() #샤디움 버튼 클릭
            M.TIME(2)
            driver.find_element(By.CLASS_NAME,"inputDefault-2F39XG.input-3xr72x").click()
            M.TIME(2)
            driver.find_element(By.CLASS_NAME,"submitButton-34IPxt.button-f2h6uQ.lookFilled-yCfaCM.colorGreen-3y-Z79.sizeMedium-2bFIHr.grow-2sR_-F").click()
            M.TIME(5)
        except:
            print("# 5 Failed")
            break

        # 6 - Verify 미작동으로 인한 pyautogui 이용한 물리적 클릭
        try:
            driver.get("https://discord.com/channels/933959587462254612/966389866017472532")#verify 이동
            M.TIME(6)
            M.CLICK_MOSUE(2000,1145) #공백클릭
            M.TIME(1)
            M.TAB(4)
            M.TIME(1)
            M.ENTER(3)
            M.TIME(2)
            M.ENTER(1)
            M.TIME(3)
        except:
            print("# 6 Failed")
            break

        # 7 - 디스코드 미션 수행 2
        try:
            driver.get("https://discord.com/channels/933959587462254612/1021737152251441244") #Faucet-1-6 이동
            M.TIME(8)
            (
            action.send_keys("/faucet").pause(2)
            .key_down(Keys.TAB).key_up(Keys.TAB)
            .send_keys(WALLET_ADDRESS)
            .key_down(Keys.ENTER).key_up(Keys.ENTER)
            .pause(2).perform() #대화창에 /faucet 입력 후 엔터
            )
            action.reset_actions() #액션 종료
        except:
            print("# 7 Failed")
            break

        # 8 - 디스코드 미션 수행 3
        try:
            driver.get("https://discord.com/channels/933959587462254612/1031497272191627284") #Faucet-2-1 이동
            M.TIME(5)
            (
            action.send_keys("/faucet").pause(2)
            .key_down(Keys.TAB).key_up(Keys.TAB)
            .send_keys(WALLET_ADDRESS)
            .key_down(Keys.ENTER).key_up(Keys.ENTER)
            .pause(2).perform() #대화창에 /faucet 입력 후 엔터
            )
            action.reset_actions() #액션 종료
            M.TIME(10)
        except:
            print("# 8 Failed")
            break
            
        # 9 - 샤디움 1.X 네트워크 Dapp
        try:
            driver.get("https://nft.shardeum.us")
            M.TIME(5)
            driver.find_element(By.CLASS_NAME,"cta-button.connect-wallet-button").click()
            M.TIME(6)
            print(driver.window_handles)
            driver.get("chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html")
            M.TIME(5)
            #driver.switch_to.window(driver.window_handles[1]) #크롬 팝업창으로 이동
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary").click() #다음
            M.TIME(2)
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary.page-container__footer-button").click() #연결
            M.TIME(5)
        except:
            print("# 9 Failed")
            break

        # 10 - 샤디움 1.X 네트워크 네이밍설정
        try:
            driver.get("https://nft.shardeum.us")
            M.TIME(5)
            driver.find_element(By.CLASS_NAME,"first-row").click() #텍스트박스 클릭
            SHARDEUM_NAME_SHM = SHARDEUM_NAME()
            print(SHARDEUM_NAME_SHM)
            (
            action.send_keys(SHARDEUM_NAME_SHM)
            .send_keys((Keys.TAB) * 2)
            .send_keys(Keys.ENTER)
            .pause(15).perform()
            )
            SHARDEUM_WS.cell(row = 1, column = (1)).value = SHARDEUM_NAME_SHM #샤디움 네임 서비스 도메인 기록
            action.reset_actions() #액션 종료
        except:
            print("# 10 Failed")
            break

        # 11 - 샤디움 1.X 10 SHM 민팅하기
        try:
            driver.get("chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html")
            M.TIME(5)
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary.page-container__footer-button").click() #보내기 확인
            M.TIME(5)
        except:
            print("# 11 Failed")
            break

        #12 - 지갑연결 팝업창 해결부분
        try:
            driver.get("https://globalswapliberty10.netlify.app/#/swap")
            M.TIME(5)
            driver.find_element(By.CLASS_NAME,"sc-bdnxRM.bhVlig.sc-kEqXSa.sc-iqAclL.hAmzBT.jJmaYB").click() #지갑연결
            M.TIME(3)
            driver.find_element(By.CLASS_NAME,"sc-oeezt.sc-hhIiOg.sc-lbVvki.eRBWbZ.iwisNA.cmzHXG").click() #메타마스크 연결
            M.TIME(5)
            driver.switch_to.window(driver.window_handles[1]) #크롬 팝업창으로 이동
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary").click() #다음
            M.TIME(2)
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary.page-container__footer-button").click() #연결
            M.TIME(5)
            driver.switch_to.window(driver.window_handles[0]) #기존 크롬창으로 이동
        except:
            print("# 12 Failed")
            break

        # 13 - 글로벌 스왑에서 샤디움 민팅 (5회 반복)
        for k in range(5):
            try:
                driver.get("https://globalswapliberty10.netlify.app/#/swap")
                M.TIME(4)
                driver.find_element(By.CLASS_NAME,"sc-clGGWX.clvOVW.token-symbol-container").click() #Token 클릭
                M.TIME(2)
                driver.find_element(By.CLASS_NAME,"sc-bdnxRM.sc-gKAaRy.sc-iCoGMd.sc-fyjqAk.fzUdiI.fHYJrX.kMthTr.cbtMgy.token-item-0xEadcbd9115Eb06698ba6e1Cd7BB4C6381f9E6729").click() #DAI 로 변경
                M.TIME(2)
                #1~15 랜덤숫자 기입(소숫점 1자리까지)
                driver.find_element(By.CLASS_NAME,"sc-fTFMiz.kclJa.token-amount-input").send_keys(M.RANDOM_NUM(5,15))#난수 숫자 기입 #From 클릭
                M.TIME(4)
                driver.find_element(By.ID,"swap-button").click() #스왑 클릭
                M.TIME(4)
                driver.find_element(By.ID,"confirm-swap-or-send").click() #컨필름 스왑 클릭
                M.TIME(8)
            except:
                print("# 13 Failed", k + 1 ,"LOOP")
                break

            # 14 - 메타마스크 팝업 처리
            try:
                driver.switch_to.window(driver.window_handles[1]) #크롬 팝업창으로 이동

                """
                #팝업창 말고 기존창에서 진행하고 싶을떄
                driver.get("chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html#confirmation")#가스비 확인을 위한 크롬 익스텐션 이동
                M.TIME(5)
                """

                driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary.page-container__footer-button").click() #가스비 확인
                M.TIME(15) #거래대기
                driver.switch_to.window(driver.window_handles[0]) #원래창으로 이동
            except:
                print("# 14 Failed")
                break

        # 15 - 샤디움 2.X 으로 네트워크 변경
        try:
            driver.get("chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html#") #네트워크 변경 위한 크롬 익스텐션 이동
            M.TIME(4) #사이트 접속대기
            driver.find_element(By.CLASS_NAME,"network-display.network-display--clickable.chip.chip--with-left-icon.chip--with-right-icon.chip--border-color-border-default.chip--background-color-undefined.chip--max-content").click() #네트워크 선택
            driver.find_element(By.CSS_SELECTOR,"[data-testid='Shardeum Liberty 2.X-network-item']").click() #샤디움 2.0으로 네트워크변경
            M.TIME(2) #네트워크 변경 대기
        except:
            print("# 15 Failed")
            break

        # 16 - 샤디움 2.X 트젝을 위한 지갑연결
        try:
            driver.get("https://app.swapped.finance/add/ETH/0x2Aaebd8534a0B377c6d7419F0226Fb01A4290C13") #사이트 이동
            M.TIME(5) #사이트 접속대기
            driver.find_element(By.ID,"connect-wallet").click() #지갑연결 클릭
            M.TIME(2)
            M.CLICK_MOSUE(1337,747)
            M.TIME(6)
        except:
            print("# 16 Failed")
            break

        # 17 - 메타마스크 팝업 / 지갑연결
        try:
            driver.switch_to.window(driver.window_handles[1]) #크롬 팝업창으로 이동
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary").click() #지갑연결 다음
            M.TIME(2) #연결 대기
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary.page-container__footer-button").click() #지갑연결 다음
            M.TIME(4) #연결 대기
            driver.switch_to.window(driver.window_handles[0]) #원래 창으로 이동
        except:
            print("# 17 Failed")
            break

        for l in range(1):
            # 18 - 샤디움 2.X 트젝하기 (1회 반복)
            try:
                driver.get("https://app.swapped.finance/swap") #스왑 사이트로 이동
                M.TIME(5)
                M.CLICK_MOSUE(1900, 700) #빈공백 아무곳이나 클릭
                action.send_keys(Keys.TAB).send_keys(M.RANDOM_NUM(5,15)).pause(7).perform()#4~15난수입력해주세요
                action.reset_actions() #액션 종료
                driver.find_element(By.ID,"swap-button").click() #스왑
                M.TIME(2)
                driver.find_element(By.ID,"confirm-swap-or-send").click() #Confirm Swap
                M.TIME(8)
            except:
                print("# 18 Failed",j + 1,"Loop")
                break

            # 19 - 메타마스크 팝업 처리
            try:
                driver.switch_to.window(driver.window_handles[1]) #크롬 팝업창으로 이동
                driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary.page-container__footer-button").click() #가스비 확인
                M.TIME(12) #거래대기
                driver.switch_to.window(driver.window_handles[0]) #원래 창으로 이동
                action.send_keys(Keys.TAB).send_keys(Keys.ENTER).pause(6).perform() #ADD 머시기
                action.reset_actions() #액션 종료
                driver.switch_to.window(driver.window_handles[1]) #크롬 팝업창으로 이동
                driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary.btn--large.page-container__footer-button").click() #추천 토큰 추가
                M.TIME(4)
                driver.switch_to.window(driver.window_handles[0]) #원래 창으로 이동
                action.send_keys((Keys.TAB) * 2).send_keys(Keys.ENTER).pause(2).perform() #Dissmiss!!!!
                action.reset_actions() #액션 종료
            except:
                print("# 19 Failed",j + 1,"Loop")
                break

        # 20 - DotSHM 연결
        # 만약 안된다면 네트워크를 다시 샤디움 1.0으로 바꿔주세요
        try:
            M.TIME(3)
            driver.get("https://dotshm.me") #사이트 이동
            M.TIME(5)
            action.send_keys(SHARDEUM_NAME_SHM).perform() #10번 과정에서 사용했던 샤디움 도메인 그대로 불러오기
            action.reset_actions() #액션 종료
            driver.find_element(By.ID,"headlessui-menu-button-:R3el6:").click() #검색
            M.TIME(3)
            driver.find_element(By.ID,"headlessui-menu-items-:r0:").click() #Register
            M.TIME(6)
        except:
            print("# 20 Failed")
            break

        # 21 - DotSHM 메타마스크 연결
        try:
            driver.switch_to.window(driver.window_handles[1]) #크롬 팝업창으로 이동
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary").click() #다음
            M.TIME(3)
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary.page-container__footer-button").click() #연결
            M.TIME(7)
            driver.switch_to.window(driver.window_handles[0]) #원래 창으로 이동
            M.CLICK_MOSUE(1081, 821) #MATIC부분 흰색영역 클릭
            action.send_keys(Keys.TAB).send_keys(Keys.ENTER).pause(4).send_keys((Keys.TAB) * 2).send_keys(Keys.ENTER).pause(10).perform()
            action.reset_actions() #액션 종료
        except:
            print("# 21 Failed")
            break

        # 22 - DotSHM 가스비 요구
        try:
            driver.switch_to.window(driver.window_handles[1]) #크롬 팝업창으로 이동
            driver.find_element(By.CLASS_NAME,"button.btn--rounded.btn-primary.page-container__footer-button").click() #다음
            driver.switch_to.window(driver.window_handles[0]) #크롬 팝업창으로 이동
            action.pause(10).send_keys((Keys.TAB) * 2).send_keys(Keys.ENTER).pause(5).perform()
            action.reset_actions() #액션 종료
        except:
            print("# 22 Failed")
            break
        
    driver.quit()
    n += 1 #다음 크롬프로필 실행