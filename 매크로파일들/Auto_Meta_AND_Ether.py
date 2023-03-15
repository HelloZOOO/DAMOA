import os
import subprocess

import pyautogui
import keyboard
import time

import openpyxl as op
import clipboard

import random
import string

os.getcwd()

#2Captcha API
API = 'a252574240a61c36a3154d05b0f9f856'

#==================이더메일 기록을 위한 기존 GMAIL정보==============
PATH_EXCEL = "C:/GMAIL_LOGIN_DATA/"
FILE_EXCLEL = "PC1_META_ETHER.xlsx"

META_ETHER_WB = op.load_workbook(PATH_EXCEL + FILE_EXCLEL)
META_ETHER_WS = META_ETHER_WB["ETHER"]
EMAIL_WS = META_ETHER_WB["EMAIL"]

print(META_ETHER_WB.sheetnames)

#==================이더메일 기록을 위한 기존 GMAIL정보==============

WIDTH, HEIGHT = pyautogui.size()
IDPW = "12121212" #메타지갑 아이디 비번

##크롬 프로필 폴더 위치 역슬래시\ 가 아닌 /슬래시로 폴더 구분
PATH_CHROME = (r"C:/CHROME_ACCOUNTS")

#폴더 내 파일 리스트화
FILE_LIST_RANGE = os.listdir(PATH_CHROME)
FILE_LIST_LNK_RANGE = [file for file in FILE_LIST_RANGE if file.endswith(".lnk")]

CHROME_NUMBER = ""

#=======================실행 / 반복 부분 ===========================
#몇번부터 실행?
ALL_RANGE_CHROME = len(FILE_LIST_LNK_RANGE)

n = int(input("몇번부터 시작? 1 ~ %d: " %ALL_RANGE_CHROME))
print(type(n),n)
n -= 1 #사용자 편의를 위해 0부터 시작이 아닌 1부터 시작


#=======================단축키 함수 정의==========================

#======================랜덤 EMAIL 생성======================

def RANDOM_EMAIL():
    RANDOM_EMAIL=20	# 문자의 개수(문자열의 크기)
    RANDOM_str = ""	# 문자열
    for i in range(RANDOM_EMAIL):
        RANDOM_str += str(random.choice(string.ascii_letters + string.digits))
    return(RANDOM_str + '@damoim.com')

#======================랜덤 EMAIL 생성======================

#파일 실행 함수 정의
def run(path):
    # 디렉토리 설정
    os.chdir(PATH_CHROME)
    # 파일 실행
    os.startfile(path)

#time 단축키
def TIME(how):
    pyautogui.sleep(how)

#tab 을 자동으로 눌러주는 함수
def TAB(how):
    for i in range(how) :
        pyautogui.hotkey('tab',interval = 0.02)

#alt + D 를 자동으로 눌러주는 함수
def ALT_D(how):
    for i in range(how) :  
        pyautogui.hotkey('alt', 'd')

#ctrl + A 를 자동으로 눌러주는 함수
def CTRL_A(how):
    for i in range(how) :  
        pyautogui.hotkey('ctrl', 'a')

#alt + F 를 자동으로 눌러주는 함수
def ALT_F4(how):
    for i in range(how) :
        pyautogui.hotkey('alt', 'f4')

#enter를 자동으로 눌러주는 함수
def ENTER(how):
    for i in range(how) :
        pyautogui.hotkey('enter')

#전체화면 단축키
def FULL_SCREEN(how):
    for i in range(how) :
        pyautogui.hotkey('f11')

#화면 깨끗하게
def CLEAR_SCREEN(how):
    for i in range(how) :
        pyautogui.hotkey('win','d')
        
#크롬 새탭
def NEW_TAB(how):
    for i in range(how) :
        pyautogui.hotkey('ctrl','t')

#크롬 탭뒤로
def BACK_TAB(how):
    for i in range(how) :
        pyautogui.hotkey('ctrl','shift','tab')

#마우스 클릭
def CLICK_MOSUE(x1, y1):
    pyautogui.click(x1, y1, button='left',clicks=1)

#문자입력
def WRITE(WORD):
    pyautogui.write(WORD)
    
#복사
def CTRL_C(how):
    pyautogui.hotkey('ctrl','c')

#클립보드카피
def CLIPBOARD(WORD):
    clipboard.copy(WORD)

#크롬강제종료
def KILL_CHROME():
    os.system('taskkill /f /im chrome.exe ') #프로세스명을 사용한 프로세스 종료

##########################################################
##########################################################
######################100회 반복 부분######################
##########################################################
##########################################################
for i in range(100):
    if(n < len(FILE_LIST_LNK_RANGE)):
        executeFile = PATH_CHROME + '/' + FILE_LIST_LNK_RANGE[n]
    else:
        break #계정이 끝에 도달했을 때 프로그램 종료
    
    CHROME_NUMBER = FILE_LIST_LNK_RANGE[n] #몇번째 크롬계정인지 확인
    print(CHROME_NUMBER) #계정 번호 출력
    META_ETHER_WS.cell(row = 1, column = (1 + i)).value = CHROME_NUMBER #엑셀에 크롬 계정 기록
    EMAIL_WS.cell(row = 1, column = (1 + i)).value = CHROME_NUMBER #엑셀에 크롬 계정 기록
    ##########################################################
    ###############메타마스크 첫 1회 설치 및 가입###############
    ##########################################################
    #1_크롬 초기화부분
    TIME(2)
    CLEAR_SCREEN(1) #화면정리
    run(executeFile)
    TIME(0.3)
    KILL_CHROME()
    TIME(0.3)
    run(executeFile)
    TIME(2)
    
    #2_메타마스크설치 크롬확장자
    ALT_D(1)
    WRITE('https://chrome.google.com/webstore/detail/metamask/nkbihfbeogaeaoehlefnkodbefgpgknn')
    ENTER(1)
    FULL_SCREEN(1) #ON
    CLICK_MOSUE(1907, 52) #프로그램 시작하면 항상 복원 창을 끄고 시작해야한다 #1907, 52
    TIME(5)

    TAB(9)
    ENTER(1) #확장프로그램 설치
    TIME(1)

    TAB(1)
    ENTER(1) #확장프로그램추가

    TIME(15) #메타마스크 들어가는데 오래걸린다    

    TAB(1)
    ENTER(1) #METAMASK 시작하기
    TIME(1.5)

    TAB(1)
    ENTER(1) #괜찮습니다
    TIME(2)

    TAB(2)
    ENTER(1) #지갑생성
    TIME(1.5)

    WRITE(IDPW)
    TAB(1)
    WRITE(IDPW)
    TAB(1)
    ENTER(1) #이용약관 동의

    TAB(2)
    ENTER(1) #계정생성
    TIME(6)

    TAB(8)
    ENTER(1) #다음
    TIME(2)

    TAB(2)
    ENTER(1) #다음에 생성
    TIME(1)
    #--------------------------------------------------------

    ##########################################################
    ###############2Captcha 1회 설치 및 가입###################
    ##########################################################
    NEW_TAB(1)
    FULL_SCREEN(1) #OFF
    ALT_D(1)
    WRITE('https://chrome.google.com/webstore/detail/captcha-solver-auto-recog/ifibfemgeogfhoebkmokieepdoobkbpo/related')
    ENTER(1)
    FULL_SCREEN(1) #ON
    TIME(5)
    
    TAB(9)
    ENTER(1) #확장프로그램 설치
    TIME(1)
    
    TAB(1)
    ENTER(1) #확장프로그램추가

    TIME(5) #2Captcha 입장 대기
    CLICK_MOSUE(50,50)
    TAB(2)
    WRITE(API)
    TAB(1)
    ENTER(1)
    TIME(3)
    CLICK_MOSUE(1135,95) #ACCESS 클릭
    TIME(2)
    NEW_TAB(1)
    TIME(1)
    KILL_CHROME() #크롬 초기화
    
    #--------------------------------------------------------
##########################################################      
##########################################################
###################### 50회 반복 부분######################
##########################################################
##########################################################
#개당 약 140초 소요 #50회 : 약 2시간
    for j in range(50):
        #1_크롬 초기화부분
        TIME(0.5)
        run(executeFile)
        TIME(0.3)
        KILL_CHROME()
        TIME(0.3)
        run(executeFile)
        TIME(2)

        #1.1_이더메일 로그아웃 확인
        ALT_D(1)
        WRITE('https://ethermail.io/webmail')
        ENTER(1)
        TIME(4)
        WRITE(IDPW)
        ENTER(1)
        TIME(8)
        ALT_D(1)
        CTRL_C(1)
        LOGOUT_RESULT = clipboard.paste()
        if '/webmail' in LOGOUT_RESULT:
            #1.2_이더메일 최종 가입 부분 재로그인 및 재로그아웃
            FULL_SCREEN(1) #ON
            TIME(1)
            CLICK_MOSUE(1907, 52) #프로그램 시작하면 항상 복원 창을 끄고 시작해야한다 #1907, 52
            TIME(0.5)
            CLICK_MOSUE(960,700) #다음
            TIME(0.6)
            TAB(1)
            EMAIL_RAND = RANDOM_EMAIL()
            EMAIL_WS.cell(row = (2 + j - 1), column = (1+i)).value = EMAIL_RAND
            WRITE(EMAIL_RAND) #랜덤 이메일 생성 함수를 제작합시다
            ENTER(1)
            TIME(0.6)
            TAB(1)
            ENTER(1)
            TIME(0.6)
            TAB(1)
            ENTER(1)
            TIME(1.5)
            CLICK_MOSUE(950,830) #암호화 키 검색
            TIME(14)
            TAB(2)
            ENTER(1) #서명
            TIME(35)
            TAB(1)
            ENTER(1)
            TIME(2.5)
            CLICK_MOSUE(880,880) #빈 화면 클릭하여 화면 좌측하단 주소 복사할 수 있게
            TIME(0.5)

            #1.3_메마 주소 엑셀에 기록 후 로그아웃
            clipboard.copy('FAIL')
            CLICK_MOSUE(143,1055) #좌측하단 주소 복사
            TIME(1.3)
            CLICK_MOSUE(100,770) #좌측중반 로그아웃
            METAMASK_RESULT = clipboard.paste()
            META_ETHER_WS.cell(row = (2 + j - 1), column = (1+i)).value = METAMASK_RESULT 
            #row : 메마주소 col : 크롬계정
            META_ETHER_WB.save(FILE_EXCLEL)
            TIME(4)
            CLEAR_SCREEN(1) #화면정리
            KILL_CHROME() #크롬종료
            
            #종료하고 다시 실행
            TIME(0.5)
            run(executeFile)
            TIME(2)
        else:
            #오류없음, 다시실행
            KILL_CHROME()
            TIME(0.3)
            run(executeFile)
            TIME(2)

        #2_이더메일입장
        ALT_D(1)
        WRITE('https://ethermail.io/?afid=636773890aa6a0fe6220c207')
        ENTER(1)
        FULL_SCREEN(1) #ON
        TIME(1)
        CLICK_MOSUE(1907, 52) #프로그램 시작하면 항상 복원 창을 끄고 시작해야한다 #1907, 52
        TIME(7)

        #3_이더메일가입
        TAB(8)
        ENTER(1)
        TIME(8.5)
        CLICK_MOSUE(833, 620) #로봇이 아닙니다 혹은 2CAPTCHA #833, 620
        TIME(55) #2CAPTCHA의 처리속도는 느리다
        CLICK_MOSUE(800, 600) #동일부분 클릭(메타마스크 클릭)
        TIME(4)

        #4_메타마스크로그인
        WRITE(IDPW)
        ENTER(1)
        
        if j <= 10:
            TIME(7) #계정이 추가되면 추가될수록 오래걸린다
        elif 10 < j <= 20:
            TIME(10)
        elif 20 < j <= 34:
            TIME(13)
        elif 34 < j <= 42:
            TIME(18)
        else:
            TIME(20)
        
        CLICK_MOSUE(1850,240) #새계정 클릭
        TIME(3)
        CTRL_A(1)
        num_chrome_str = str(i + 1) #i번 크롬
        num_meta_str = str(j + 1) #j번 지갑
        WRITE('A' + '_' + num_chrome_str.zfill(3) + '_' + num_meta_str.zfill(2)) #ACCOUNT_크롬계정번호_지갑번호
        ENTER(1)

        if j <= 10:
            TIME(8) #계정이 추가되면 추가될수록 오래걸린다
        elif 10 < j <= 20:
            TIME(11)
        elif 20 < j <= 34:
            TIME(14)
        elif 34 < j <= 42:
            TIME(18)
        else:
            TIME(22)

        CLICK_MOSUE(1832,570) #다음
        TIME(0.5)
        CLICK_MOSUE(1832,570) #연결
        TIME(4)
        CLICK_MOSUE(1832,570) #서명
        TIME(45) #서명하고 이더메일 들어가는게 느리다

        #5_서명 제대로 되었나 확인
        FULL_SCREEN(1) #OFF
        ALT_D(1)
        WRITE('https://ethermail.io/webmail')
        ENTER(1)
        TIME(2)
        ALT_D(1)
        TIME(0.2)
        CTRL_C(1)
        TIME(0.2)
        FULL_SCREEN(1) #ON
        ROGIN_RESULT = clipboard.paste()
        
        if '/webmail' in ROGIN_RESULT:
        
            #6_이더메일 최종 가입 부분
            TIME(1)
            CLICK_MOSUE(960,700) #다음
            TIME(0.6)
            TAB(1)
            EMAIL_RAND = RANDOM_EMAIL()
            EMAIL_WS.cell(row = (2 + j), column = (1+i)).value = EMAIL_RAND
            WRITE(EMAIL_RAND) #랜덤 이메일 생성 함수를 제작합시다
            ENTER(1)
            TIME(0.6)
            TAB(1)
            ENTER(1)
            TIME(0.6)
            TAB(1)
            ENTER(1)
            TIME(1.5)
            CLICK_MOSUE(950,830) #암호화 키 검색
            TIME(14)
            TAB(2)
            ENTER(1) #서명
            TIME(18)
            TAB(1)
            ENTER(1)
            TIME(2.5)
            CLICK_MOSUE(880,880) #빈 화면 클릭하여 화면 좌측하단 주소 복사할 수 있게
            TIME(0.5)

            #7_메마 주소 엑셀에 기록 후 로그아웃
            clipboard.copy('FAIL')
            CLICK_MOSUE(143,1055) #좌측하단 주소 복사
            TIME(1.3)
            CLICK_MOSUE(100,770) #좌측중반 로그아웃
            METAMASK_RESULT = clipboard.paste()
            META_ETHER_WS.cell(row = (2 + j), column = (1+i)).value = METAMASK_RESULT 
            #row : 메마주소 col : 크롬계정
            META_ETHER_WB.save(FILE_EXCLEL)
            TIME(4)
            CLEAR_SCREEN(1) #화면정리
            KILL_CHROME() #크롬종료
        else:
            METAMASK_RESULT = 'FAIL'
            META_ETHER_WS.cell(row = (2 + j), column = (1+i)).value = METAMASK_RESULT
            META_ETHER_WB.save(FILE_EXCLEL)
            CLEAR_SCREEN(1) #화면정리
            KILL_CHROME() #크롬종료
        
        print("이더메일",j,"회 생성중")
    
    n += 1 #다음 크롬프로필
    print("크롬계정",i,"회 가동중")
print("프로그램 종료")