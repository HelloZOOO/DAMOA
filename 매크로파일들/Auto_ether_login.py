import os
import subprocess

import pyautogui
import keyboard
import time

import openpyxl as op
import clipboard

os.getcwd()

#==================이더메일 가입을 위한 기존 GMAIL정보==============
PATH_EXCEL = "C:/GMAIL_LOGIN_DATA/"
FILE_EXCLEL = "PC2_ALL_ACCOUNT.xlsx"

GMAIL_WB = op.load_workbook(PATH_EXCEL + FILE_EXCLEL)
GMAIL_WS_ACCOUNT = GMAIL_WB["CHECK"]
GMAIL_WS_ETHER = GMAIL_WB["ETHER"]

MAX_ROW_ACCOUNT = GMAIL_WS_ACCOUNT.max_row

print(GMAIL_WB.sheetnames)

#==================이더메일 가입을 위한 기존 GMAIL정보==============

WIDTH, HEIGHT = pyautogui.size()
IDPW = "12121212" #메타지갑 아이디 비번

##크롬 프로필 폴더 위치 역슬래시\ 가 아닌 /슬래시로 폴더 구분
PATH_CHROME = (r"C:/CHROME_ACCOUNTS")

#폴더 내 파일 리스트화
FILE_LIST_RANGE = os.listdir(PATH_CHROME)
FILE_LIST_LNK_RANGE = [file for file in FILE_LIST_RANGE if file.endswith(".lnk")]

CHROME_NUMBER = ""

#=======================실행 / 반복 부분 ===========================
#몇번부터 실행?
ALL_RANGE_CHROME = len(FILE_LIST_LNK_RANGE)
ALL_RANGE_ROW = MAX_ROW_ACCOUNT

n = int(input("몇번부터 시작? 1 ~ %d: " %ALL_RANGE_ROW))
print(type(n),n)
n -= 1 #사용자 편의를 위해 0부터 시작이 아닌 1부터 시작

#파일 실행 함수 정의
def run(path):
    # 디렉토리 설정
    os.chdir(PATH_CHROME)
    # 파일 실행
    os.startfile(path)


#=======================단축키 함수 정의==========================

#크롬실행시 기본 설정 (페이지 복원메시지 끄기)
def START_CHROME():
    pyautogui.sleep(0.5)
    pyautogui.hotkey('f11')
    pyautogui.click(1861, 203, button='left',clicks=1)
    pyautogui.hotkey('f11')

#time 단축키
def TIME(how):
    pyautogui.sleep(how)

#tab 을 자동으로 눌러주는 함수
def TAB(how):
    for i in range(how) :
        pyautogui.hotkey('tab',interval = 0.02)

#alt + D 를 자동으로 눌러주는 함수
def ALT_D(how):
    for i in range(how) :  
        pyautogui.hotkey('alt', 'd')

#alt + F 를 자동으로 눌러주는 함수
def ALT_F4(how):
    for i in range(how) :
        pyautogui.hotkey('alt', 'f4')

#enter를 자동으로 눌러주는 함수
def ENTER(how):
    for i in range(how) :
        pyautogui.hotkey('enter')

#전체화면 단축키
def FULL_SCREEN(how):
    for i in range(how) :
        pyautogui.hotkey('f11')

#화면 깨끗하게
def CLEAR_SCREEN(how):
    for i in range(how) :
        pyautogui.hotkey('win','d')
        
#크롬 새탭
def NEW_TAB(how):
    for i in range(how) :
        pyautogui.hotkey('ctrl','t')

#크롬 탭뒤로
def BACK_TAB(how):
    for i in range(how) :
        pyautogui.hotkey('ctrl','shift','tab')

#마우스 클릭
def CLICK_MOSUE(x1, y1):
    pyautogui.click(x1, y1, button='left',clicks=1)

#문자입력
def WRITE(WORD):
    pyautogui.write(WORD)
    
#복사
def CTRL_C(how):
    pyautogui.hotkey('ctrl','c')

#클립보드카피
def CLIPBOARD(WORD):
    clipboard.copy(WORD)

#크롬강제종료
def KILL_CHROME():
    os.system('taskkill /f /im chrome.exe ') #프로세스명을 사용한 프로세스 종료



#=======================단축키 함수 부분 끝==========================

for i in range(MAX_ROW_ACCOUNT):
    #특정 파일 링크 - 파일에는 한글 공백 특수기호가 없어야합니다
    if(n < len(FILE_LIST_LNK_RANGE)):
        executeFile = PATH_CHROME + '/' + FILE_LIST_LNK_RANGE[n]
    else:
        break

    #몇번째 크롬계정인지 확인
    CHROME_NUMBER = FILE_LIST_LNK_RANGE[n]
    print(CHROME_NUMBER) #계정 번호 출력

    #==========================엑셀에서 gmail로그인 계정 불러오기==========================
    data1_ID = GMAIL_WS_ACCOUNT.cell(row = (1 + n) , column = 2).value
    print(data1_ID)
    
    EMAIL = data1_ID
    
    if type(EMAIL) is int:
        print("정수형을 문자열로 변환합니다")
        EMAIL = str(data1_ID)
    
    
    if type(EMAIL) is float:
        print("정수형을 문자열로 변환합니다")
        EMAIL = str(data1_ID)
    #==========================엑셀에서 gmail로그인 계정 불러오기==========================
    
    TIME(1)
    CLEAR_SCREEN(1)
    run(executeFile) #크롬실행
    
    #============================이더메일 계정 생성==========================

    TIME(8) #메타마스크 뜰때까지 대기
    NEW_TAB(1)
    
    TIME(0.7)
    ALT_D(1)
    WRITE('https://ethermail.io/?afid=636773890aa6a0fe6220c207')
    ENTER(1)
    FULL_SCREEN(1)
    TIME(1)
    CLICK_MOSUE(1907, 52) #프로그램 시작하면 항상 복원 창을 끄고 시작해야한다 #1907, 52
    TIME(7)
    
    #이더메일 가입
    TAB(8)
    ENTER(1)
    TIME(5)
    CLICK_MOSUE((WIDTH * 43.54 / 100), (HEIGHT * 55.27 / 100)) #로봇이 아닙니다 #836,597
    FULL_SCREEN(1)
    TIME(8)
    
    #캡차우회부분
    TAB(2)
    ENTER(1)#캡차 버스터 엔터
    TIME(3)
    FULL_SCREEN(1)
    TIME(2)
    #CLICK_MOUSE(1071, 707) #필요하면 활성화 Vertify! 부분임
    #TIME(3)
    CLICK_MOSUE((WIDTH * 43.54 / 100), (HEIGHT * 55.27 / 100)) #메타마스크 클릭 #836,597
    TIME(4)
    
    #메타마스크 로그인
    WRITE(IDPW)
    ENTER(1)
    TIME(3)
    FULL_SCREEN(1)
    CLICK_MOSUE((WIDTH * 43.54 / 100), (HEIGHT * 55.27 / 100)) #TAB키 작동을 위한 빈칸클릭
    TAB(2)
    ENTER(1) #다음
    TIME(1)
    TAB(2)
    ENTER(1) #다음
    TIME(15)
    
    #서명 요청
    FULL_SCREEN(1)
    TAB(2)
    ENTER(1) #서명
    TIME(35) #서명하고 대기 오래걸림
    
    #이더메일부분
    CLICK_MOSUE(WIDTH/2, (HEIGHT * 64.8 / 100)) #가입하고 다음 #Y:700
    TIME(3)
    TAB(1)
    TIME(0.3)
    WRITE(EMAIL)
    ENTER(1)
    TIME(1)
    CLICK_MOSUE(WIDTH/2, (HEIGHT * 79.63 / 100)) #다음 #Y:860
    TIME(1)
    CLICK_MOSUE(WIDTH/2, (HEIGHT * 79.63 / 100)) #Go to inbox! #Y:860
    TIME(1)
    CLICK_MOSUE(WIDTH/2, (HEIGHT * 75.92 / 100)) #Retrieve Encryption Keys #Y:820
    TIME(6)
    
    #서명 요청
    FULL_SCREEN(1)
    TAB(2)
    ENTER(1) #서명
    TIME(6)
    CLICK_MOSUE(WIDTH/2, (HEIGHT * 74.07 / 100)) #가입 끝 #Y:800
    TIME(3)
    
    #크롬 프로세스 종료하고 다시 실행시켜서 메일로 이동하기
    #마무리 가입 확인 여부 코드
    KILL_CHROME() #크롬종료
    run(executeFile) #크롬실행
    TIME(0.5)
    ALT_D(1)
    WRITE('https://ethermail.io/webmail/')
    TIME(1)
    ENTER(1)
    CLIPBOARD("FAIL")
    TIME(5) #메타마스크 뜰때까지 대기
    NEW_TAB(1)
    BACK_TAB(1)

    #현재 창 주소 복사
    ALT_D(1)
    CTRL_C(1)
    ETHER_RESULT = clipboard.paste()
    CHECK_RESULT = ETHER_RESULT
    if 'https://ethermail.io/webmail/' in CHECK_RESULT:
        CHECK_RESULT = "O"
    else:
        CHECK_RESULT = "X"
        
    
    #==========================엑셀에 클립보드 내용 저장==========================
    GMAIL_WS_ETHER.cell(row = (1 + n), column = 1).value = CHROME_NUMBER
    GMAIL_WS_ETHER.cell(row = (1 + n), column = 2).value = EMAIL
    GMAIL_WS_ETHER.cell(row = (1 + n), column = 3).value = CHECK_RESULT
    GMAIL_WS_ETHER.cell(row = (1 + n), column = 4).value = ETHER_RESULT
    print ("%s 계정 %s에 저장되었습니다" % (CHROME_NUMBER, FILE_EXCLEL),ETHER_RESULT)
    GMAIL_WB.save(FILE_EXCLEL)
    #==========================엑셀에 클립보드 내용 저장==========================

    KILL_CHROME() #크롬종료
    CLEAR_SCREEN(1) #화면정리
    TIME(1.5)
    #============================이더메일 계정 생성==========================
    #한번 더 반복
    n += 1

GMAIL_WB.save(FILE_EXCLEL)
print("이더메일 자동가입 종료")