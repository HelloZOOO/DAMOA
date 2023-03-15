import os

import pyautogui
import keyboard
import time

import openpyxl as op
import clipboard

os.getcwd()
WIDTH, HEIGHT = pyautogui.size()

#=================엑셀 시트 임포트 설정===================#
PATH_GMAIL_CHECK = "C:/GMAIL_LOGIN_DATA/"
FILE_GMAIL_CHECK = "PC1_ALL_ACOUNT.xlsx"

GMAIL_CHECK_WB = op.load_workbook(PATH_GMAIL_CHECK + FILE_GMAIL_CHECK)
GMAIL_CHECK_WS = GMAIL_CHECK_WB.active
#=================엑셀 시트 임포트 설정===================#

#파일 실행 함수 정의
def RUN(path):
    # 디렉토리 설정
    os.chdir(PATH_CHROME)
    # 파일 실행
    os.startfile(path)

#============함수설정부분============#
#TAB 을 자동으로 눌러주는 함수
def TAB(how):
    for i in range(how) :
        pyautogui.hotkey('tab',interval=0.02)

#alt + D 를 자동으로 눌러주는 함수
def ALT_D(how):
    for i in range(how) :   
        pyautogui.hotkey('alt', 'd')

#alt + F 를 자동으로 눌러주는 함수
def ALT_F4(how):
    for i in range(how) :
        pyautogui.hotkey('alt', 'f4')

#ENTER를 자동으로 눌러주는 함수
def ENTER(how):
    for i in range(how) :
        pyautogui.hotkey('enter')

#ctrl + t를 자동으로 눌러주는 함수(크롬 새탭)
def CTRL_T(how):
    for i in range(how) :
        pyautogui.hotkey('ctrl', 't')

#F11를 자동으로 눌러주는 함수
def FULLSCREEN_F11(how):
    for i in range(how) :
        pyautogui.hotkey('f11')

#바탕화면나가기
def WIN_D(how):
    for i in range(how) :
        pyautogui.hotkey('win','d')

#마우스 클릭
def CLICK_MOSUE(x1, y1):
    pyautogui.click(x1, y1, button='left',clicks=1)

#마우스 DRAG하고 복사해주는 함수
def CLICK_COPY(x1, y1):
    pyautogui.click(x1, y1, button='left',clicks=2)
    time.sleep(0.2)
    pyautogui.click(x1, y1, button='left',clicks=2)
    pyautogui.hotkey('ctrl','c')

#============함수설정부분============#


#=================파일 리스트 실행===================

#폴더 위치 역슬래시\ 가 아닌 /슬래시로 폴더 구분
PATH_CHROME = (r"C:/CHROME_ACCOUNTS")

#폴더 내 파일 리스트화
FILE_LIST_RANGE = os.listdir(PATH_CHROME)
FILE_LIST_LNK_RANGE = [file for file in FILE_LIST_RANGE if file.endswith(".lnk")]

#몇번부터 실행?
ALL_RANGE = len(FILE_LIST_LNK_RANGE)
n = int(input("몇번부터 시작? 1 ~ %d : " %ALL_RANGE))
print(type(n),n)
n -= 1 #사용자 편의를 위해 0부터 시작이 아닌 1부터 시작

for i in FILE_LIST_LNK_RANGE:
    if(n < len(FILE_LIST_LNK_RANGE)):
        executeFile = PATH_CHROME + '/' + FILE_LIST_LNK_RANGE[n]
    
        #특정 파일 링크 - 파일에는 한글 공백 특수기호가 없어야합니다
        CHROME_ACOUNT = FILE_LIST_LNK_RANGE[n]
        print(CHROME_ACOUNT)

        RUN(executeFile) #프로그램 실행
        
        #=================파일 리스트 종료===================
        
        #=================키보드 매크로 설정===================
        
        time.sleep(2.5) #잠시 대기시간을 주자
        
        CTRL_T(1)
        time.sleep(0.2)
        FULLSCREEN_F11(1)
        
        time.sleep(6) #메타마스크화면뜨는거 기다리기...
        CLICK_MOSUE(315, 640) #메타마스크 화면뜨는거 빡치네
        
        TAB(4)
        ENTER(1)
        
        time.sleep(0.5)
        CLICK_COPY((WIDTH * 81.5 / 100), 214)
        
        time.sleep(.5)
        ALT_F4(1)
        WIN_D(1)
        time.sleep(1)
        #==========================엑셀에 클립보드 내용 저장==========================
        CHECK_RESULT = clipboard.paste()

        GMAIL_CHECK_WS.cell(row = (1 + n), column = 1).value = CHECK_RESULT
        GMAIL_CHECK_WS.cell(row = (1 + n), column = 2).value = CHROME_ACOUNT #파일 이름 저장
        print ("%s에 저장되었습니다" % (PATH_CHROME + "/" + FILE_GMAIL_CHECK),CHECK_RESULT)
        GMAIL_CHECK_WB.save(FILE_GMAIL_CHECK) #구글계정 엑셀에 저장
        #==========================엑셀에 클립보드 내용 저장==========================
        
        #=================키보드 매크로 종료===================
        n += 1 #다음 파일 실행 / 엑셀 다음 행렬
    else:
        break
path = GMAIL_CHECK_WB
print("엑셀 저장 위치")
print(PATH_CHROME + "/" + FILE_GMAIL_CHECK)
print("GMAIL 메일확인 프로그램 종료")