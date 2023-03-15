import os

import pyautogui
import keyboard
import time

import openpyxl as op
import clipboard

import Main as M

os.getcwd()

#==================메타지갑 저장 위한 엑셀 파일 폴더설정==============
PATH_EXCEL = "C:/DAMOA_DATA/"
FILE_EXCLEL = "PC6_METAMASK_ADDRESS_RE.xlsx"

EXELFILENAME = FILE_EXCLEL

META_WB = op.load_workbook(PATH_EXCEL + FILE_EXCLEL)
META_WS = META_WB.active
#==================메타지갑 저장 위한 엑셀 파일 폴더설정==============

WIDTH, HEIGHT = pyautogui.size()
IDPW = "12121212" #메타지갑 아이디 비번

##크롬 프로필 폴더 위치 역슬래시\ 가 아닌 /슬래시로 폴더 구분
PATH_CHROME = (r"C:/CHROME_ACCOUNTS")

#폴더 내 파일 리스트화
FILE_LIST_RANGE = os.listdir(PATH_CHROME)
FILE_LIST_LNK_RANGE = [file for file in FILE_LIST_RANGE if file.endswith(".lnk")]

CHROME_NUMBER = ""

#파일 실행 함수 정의
def run(path):
    # 디렉토리 설정
    os.chdir(PATH_CHROME)
    # 파일 실행
    os.startfile(path)

#=======================실행 / 반복 부분 ===========================
#몇번부터 실행?
ALL_RANGE = len(FILE_LIST_LNK_RANGE)
n = int(input("몇번부터 시작? 1 ~ %d: " %ALL_RANGE))
print(type(n),n)
n -= 1 #사용자 편의를 위해 0부터 시작이 아닌 1부터 시작

for i in FILE_LIST_LNK_RANGE:
    #특정 파일 링크 - 파일에는 한글 공백 특수기호가 없어야합니다
    if(n < len(FILE_LIST_LNK_RANGE)):
        executeFile = PATH_CHROME + '/' + FILE_LIST_LNK_RANGE[n]
    else:
        break

    #몇번째 크롬계정인지 확인
    CHROME_NUMBER = FILE_LIST_LNK_RANGE[n]
    print(CHROME_NUMBER) #계정 번호 출력

    run(executeFile) #크롬실행
    
    #============================메타마스크 계정 생성==========================

    M.TIME(1.5) #잠시 대기시간을 주자
    M.NEW_TAB(1)
    M.TIME(1)
    M.ALT_D(1)
    M.WRITE('chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html')
    M.ENTER(1)
    M.TIME(3)
    
    #메타마스크 로그인
    M.WRITE(IDPW)
    M.ENTER(1)
    M.TIME(3)
    
    #새계정생성
    M.ALT_D(1)
    M.WRITE('chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/HOME.HTML#new-account')
    M.ENTER(1)
    M.FULL_SCREEN(1) #ON
    M.TIME(3)
    M.TAB(2)
    M.ENTER(1)
    M.TIME(3)
    
    #첫 로그인시 팝업 클릭 후 메타마스크 재접속
    M.CLICK_MOSUE((WIDTH/2), 100)
    M.TIME(0.2)
    M.CLICK_MOSUE((WIDTH/2), 100)
    M.TIME(0.2)
    M.CLICK_MOSUE((WIDTH/2), 680)
    M.TIME(0.2)
    M.CLICK_MOSUE(10, 10)
    M.TIME(0.2)
    M.TAB(2)
    M.ENTER(1)
    M.TAB(3)
    M.ENTER(1)
    M.TIME(1)
    
    #클립보드 복사 클릭
    M.CLICK_MOSUE((WIDTH/2), 100)
    M.TIME(0.2)
    M.CLICK_MOSUE((WIDTH/2), 100)
    M.TIME(0.3)

    #==========================엑셀에 클립보드 내용 저장==========================
    META_RESULT = clipboard.paste()

    META_WS.cell(row = (1 + n), column = 1).value = META_RESULT
    META_WS.cell(row = (1 + n), column = 2).value = CHROME_NUMBER #파일 이름 저장
    print ("%s 계정 %s에 저장되었습니다" % (CHROME_NUMBER, FILE_EXCLEL),META_RESULT)
    META_WB.save(EXELFILENAME) #메타지갑 엑셀 저장
    #==========================엑셀에 클립보드 내용 저장==========================

    M.ALT_F4(1) #크롬종료
    #pyautogui.press('down')

    #============================메타마스크 계정 생성==========================


    #한번 더 반복
    n += 1

META_WB.save(EXELFILENAME) #메타지갑 엑셀 저장
print("메타마스크 자동생성 파일 종료")