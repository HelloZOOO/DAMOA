import os

import pyautogui
import keyboard
import time

import openpyxl as op
import clipboard

import Main as M

#해당부분 폴더에 맞춰서 크롬 프로필 설정해주세요
PUBLIC_PATH = "C:/CHROME_ACCOUNTS_2"

#IP구역 분할을위한 변수설정
#THIS_PC부분 수정해주세요
PC0 = 0
PC1 = 1000
PC2 = 2000
PC3 = 3000
PC6 = 4000
THIS_PC = PC1

WIDTH, HEIGHT = pyautogui.size()
#==================IP가 저장된 엑셀파일 열기==============
PATH_EXCEL = PUBLIC_PATH + "/"
FILE_EXCLEL = "EliteIP_0315.xlsx"

EXELFILENAME = FILE_EXCLEL

IP_WB = op.load_workbook(PATH_EXCEL + FILE_EXCLEL)
IP_WS = IP_WB.active
#==================IP가 저장된 엑셀파일 열기==============

##크롬 프로필 폴더 위치
PATH_CHROME = (PUBLIC_PATH)

#폴더 내 파일 리스트화
FILE_LIST_RANGE = os.listdir(PATH_CHROME)
FILE_LIST_LNK_RANGE = [file for file in FILE_LIST_RANGE if file.endswith(".lnk")]

CHROME_NUMBER = ""

#파일 실행 함수 정의
def run(path):
    # 디렉토리 설정
    os.chdir(PATH_CHROME)
    # 파일 실행
    os.startfile(path)

#=======================실행 / 반복 부분 ===========================
#몇번부터 실행?
ALL_RANGE = len(FILE_LIST_LNK_RANGE)
n = int(input("몇번부터 시작? 1 ~ %d: " %ALL_RANGE))
print(type(n),n)
n -= 1 #사용자 편의를 위해 0부터 시작이 아닌 1부터 시작

for i in FILE_LIST_LNK_RANGE:
    #특정 파일 링크 - 파일에는 한글 공백 특수기호가 없어야합니다
    if(n < len(FILE_LIST_LNK_RANGE)):
        executeFile = PATH_CHROME + '/' + FILE_LIST_LNK_RANGE[n]
    else:
        break

    #몇번째 크롬계정인지 확인
    CHROME_NUMBER = FILE_LIST_LNK_RANGE[n]
    print(CHROME_NUMBER) #계정 번호 출력

    run(executeFile) #크롬실행
    #============================IP프록시 설정==========================

    M.TIME(1.5) #잠시 대기시간을 주자
    M.ALT_D(1)
    clipboard.copy("https://chrome.google.com/webstore/detail/proxy-switchyomega/padekgcemlokbadohgkifijomclgjgif/related")
    pyautogui.hotkey('ctrl','v')
    M.ENTER(1)
    #M.FULL_SCREEN(1) #ON
    M.TIME(5)
    
    M.TAB(8) #크롬에 추가로 이동
    M.ENTER(1)
    M.TIME(1)
    
    M.TAB(1) #확장 프로그램 추가
    M.ENTER(1)
    #M.FULL_SCREEN(1) #OFF
    M.TIME(5) #설치완료 후 팝업까지 대기
    
    M.NEW_TAB(1)
    clipboard.copy("chrome-extension://padekgcemlokbadohgkifijomclgjgif/options.html#!/profile/proxy")
    pyautogui.hotkey('ctrl','v')
    M.ENTER(1)
    M.FULL_SCREEN(1) #전체화면 ON
    M.TIME(1)
    
    #IP입력
    M.TAB(11)
    IP = IP_WS.cell(row = (1 + THIS_PC + n), column = 1).value
    M.CLIPBOARD(IP) #클립보드에 아이피 복사
    pyautogui.hotkey('ctrl','v')
    
    #PORT입력
    M.TAB(1)
    PORT = IP_WS.cell(row = (1 + THIS_PC + n), column = 2).value
    M.CLIPBOARD(PORT) #클립보드에 포트번호 복사
    pyautogui.hotkey('ctrl','v')
    
    M.CLICK_MOSUE(100,430) #Apply 클릭
    M.TIME(2)
    
    M.ALT_F4(1) # 크롬 종료
    
    #한번 더 반복
    n += 1

print("IP설정 종료")