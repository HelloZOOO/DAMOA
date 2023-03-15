import os
import subprocess

import pyautogui
import keyboard
import time

import openpyxl as op
import clipboard

os.getcwd()
WIDTH, HEIGHT = pyautogui.size()

#=================엑셀 시트 임포트 설정===================

PATH_GMAIL = "C:/GMAIL_LOGIN_DATA/"
FILE_GMAIL = "PC4_ALL_ACCOUNT.xlsx"

GMAIL_WB = op.load_workbook(PATH_GMAIL + FILE_GMAIL)
GMAIL_WS_ALLACCOUNT = GMAIL_WB["ALLACCOUNT"]
GMAIL_WS_NEWACCOUNT = GMAIL_WB["CHECK"]

MAX_ROW_ACCOUNT = GMAIL_WS_ALLACCOUNT.max_row

#=================엑셀 시트 임포트 설정===================

#파일 실행 함수 정의
def run(path):
    # 디렉토리 설정
    os.chdir(PATH_CHROME)
    # 파일 실행
    os.startfile(path)

#=================파일 리스트 실행===================

#크롬실행시 기본 설정 (페이지 복원메시지 끄기)
def START_CHROME():
    pyautogui.sleep(0.5)
    pyautogui.hotkey('f11')
    pyautogui.click(1861, 203, button='left',clicks=1)
    pyautogui.hotkey('f11')

#time 단축키
def TIME(how):
    pyautogui.sleep(how)

#tab 을 자동으로 눌러주는 함수
def TAB(how):
    for i in range(how) :
        pyautogui.hotkey('tab',interval = 0.02)

#alt + D 를 자동으로 눌러주는 함수
def ALT_D(how):
    for i in range(how) :  
        pyautogui.hotkey('alt', 'd')

#alt + F 를 자동으로 눌러주는 함수
def ALT_F4(how):
    for i in range(how) :
        pyautogui.hotkey('alt', 'f4')

#ENTER를 자동으로 눌러주는 함수
def ENTER(how):
    for i in range(how) :
        pyautogui.hotkey('ENTER')

#전체화면 단축키
def FULL_SCREEN(how):
    for i in range(how) :
        pyautogui.hotkey('f11')

#화면 깨끗하게
def CLEAR_SCREEN(how):
    for i in range(how) :
        pyautogui.hotkey('win','d')
        
#크롬 새탭
def NEW_TAB(how):
    for i in range(how) :
        pyautogui.hotkey('ctrl','t')

#크롬 탭뒤로
def BACK_TAB(how):
    for i in range(how) :
        pyautogui.hotkey('ctrl','shift','tab')

#마우스 클릭
def CLICK_MOSUE(x1, y1):
    pyautogui.click(x1, y1, button='left',clicks=1)

#마우스 DRAG하고 복사해주는 함수
def CLICK_COPY(x1, y1):
    pyautogui.click(x1, y1, button='left',clicks=2)
    time.sleep(0.2)
    pyautogui.click(x1, y1, button='left',clicks=2)
    pyautogui.hotkey('ctrl','c')

#문자입력
def WRITE(WORD):
    pyautogui.write(WORD)
    
#복사
def CTRL_C(how):
    pyautogui.hotkey('ctrl','c')

#클립보드카피
def CLIPBOARD(WORD):
    clipboard.copy(WORD)

#크롬강제종료
def KILL_CHROME():
    os.system('taskkill /f /im chrome.exe ') #프로세스명을 사용한 프로세스 종료

def CHECK_EMAIL(clipboard):
    if '@gmail.com' in clipboard:
        return 1 #로그인이 되었으면 1
    else:
        return 0 #로그인이 안되었으면 0

#폴더 위치 역슬래시\ 가 아닌 /슬래시로 폴더 구분
PATH_CHROME = (r"C:/CHROME_ACCOUNTS")

#폴더 내 파일 리스트화
FILE_LIST_RANGE = os.listdir(PATH_CHROME)
FILE_LIST_LNK_RANGE = [file for file in FILE_LIST_RANGE if file.endswith(".lnk")]

#몇번부터 실행?
ALL_RANGE = len(FILE_LIST_LNK_RANGE)
ALL_RANGE_ROW = MAX_ROW_ACCOUNT
n = int(input("몇번부터 시작? 1 ~ %d : " %ALL_RANGE_ROW))
print(type(n),n)
n -= 1 #사용자 편의를 위해 0부터 시작이 아닌 1부터 시작

for i in range(ALL_RANGE_ROW):
    #특정 파일 링크 - 파일에는 한글 공백 특수기호가 없어야합니다
    executeFile = PATH_CHROME + '/' + FILE_LIST_LNK_RANGE[n]
    print(executeFile)
    
    #몇번째 크롬계정인지 확인
    CHROME_NUMBER = FILE_LIST_LNK_RANGE[n]
    print(CHROME_NUMBER) #계정 번호 출력
    
    #=================파일 리스트 종료===================
    
    data1_ID = GMAIL_WS_ALLACCOUNT.cell(row = (1 + n) , column = 2).value
    data2_PW = GMAIL_WS_ALLACCOUNT.cell(row = (1 + n) , column = 3).value
    data3_REC_ID= GMAIL_WS_ALLACCOUNT.cell(row = (1 + n) , column = 4).value
    
    EMAIL = data1_ID
    PW = data2_PW
    R_EMAIL = data3_REC_ID
    
    print(EMAIL,PW,R_EMAIL)
    
    if type(PW) is int:
        print("정수형을 문자열로 변환합니다")
        PW = str(data2_PW)
    
    
    if type(PW) is float:
        print("정수형을 문자열로 변환합니다")
        PW = str(data2_PW)
    
    
    #=================키보드 매크로 설정===================
    
    run(executeFile) #프로그램 실행
    
    time.sleep(2) #잠시 대기시간을 주자
    
    FULL_SCREEN(1)
    TIME(1)
    CLICK_MOSUE(1907, 52)
    TIME(3)
    FULL_SCREEN(1)
    NEW_TAB(1)
    TIME(0.3)
    
    ALT_D(1)
    WRITE('https://accounts.google.com/')
    ENTER(1)
    TIME(3)
    
    ALT_D(1)
    TIME(0.2)
    CTRL_C(1)
    CHECK_CLIPBOARD = clipboard.paste()
    
    if 'myaccount.google.com/' in CHECK_CLIPBOARD:
        
        #이메일 체크부분
        NEW_TAB(1)
        TIME(1)
        FULL_SCREEN(1)
        TIME(1)
        
        TAB(4)
        ENTER(1)
        
        TIME(0.5)
        CLICK_COPY((WIDTH * 81.5 / 100), 214)
        TIME(0.5)
        
        CHECK_GMAIL = clipboard.paste()
        CHECK_ID = CHECK_EMAIL(CHECK_GMAIL) #로그인 여부 반환 1: 성공 0: 실패
        
        #==========================엑셀에 클립보드 내용 저장==========================
        CHECK_RESULT = clipboard.paste()
        
        GMAIL_WS_NEWACCOUNT.cell(row = (1 + n), column = 1).value = CHECK_ID
        GMAIL_WS_NEWACCOUNT.cell(row = (1 + n), column = 2).value = EMAIL
        GMAIL_WS_NEWACCOUNT.cell(row = (1 + n), column = 3).value = PW
        GMAIL_WS_NEWACCOUNT.cell(row = (1 + n), column = 4).value = R_EMAIL
        GMAIL_WS_NEWACCOUNT.cell(row = (1 + n), column = 5).value = CHROME_NUMBER
        GMAIL_WS_NEWACCOUNT.cell(row = (1 + n), column = 6).value = CHECK_GMAIL
        print("이미 로그인 됨",EMAIL,PW,R_EMAIL,CHROME_NUMBER)
        
        KILL_CHROME()
        n += 1 #다음 파일 실행 / 엑셀 다음 행렬
        GMAIL_WB.save(FILE_GMAIL)
        
    elif 'accounts.google.com/' or 'www.google.com/intl/' in CHECK_CLIPBOARD:
        ALT_D(1)
        WRITE('https://accounts.google.com/')
        ENTER(1)
        
        TIME(2)
        
        WRITE(EMAIL)
        ENTER(1)
        TIME(2)
        
        WRITE(PW)
        ENTER(1)
        TIME(5)
        
        TAB(5)
        ENTER(1)
        TIME(2)
        
        WRITE(R_EMAIL)
        ENTER(1)
        TIME(5)
        
        #이메일 체크부분
        NEW_TAB(1)
        TIME(1)
        FULL_SCREEN(1)
        TIME(1)
        
        TAB(4)
        ENTER(1)
        
        TIME(0.5)
        CLICK_COPY((WIDTH * 81.5 / 100), 214)
        
        CHECK_GMAIL = clipboard.paste()
        CHECK_ID = CHECK_EMAIL(CHECK_GMAIL) #로그인 여부 반환 1: 성공 0: 실패
        
        TIME(0.5)
        
        KILL_CHROME()
        
        TIME(1)
        
        #=================키보드 매크로 종료===================
        
        GMAIL_WS_NEWACCOUNT.cell(row = (1 + n), column = 1).value = CHECK_ID
        GMAIL_WS_NEWACCOUNT.cell(row = (1 + n), column = 2).value = EMAIL
        GMAIL_WS_NEWACCOUNT.cell(row = (1 + n), column = 3).value = PW
        GMAIL_WS_NEWACCOUNT.cell(row = (1 + n), column = 4).value = R_EMAIL
        GMAIL_WS_NEWACCOUNT.cell(row = (1 + n), column = 5).value = CHROME_NUMBER
        GMAIL_WS_NEWACCOUNT.cell(row = (1 + n), column = 6).value = CHECK_GMAIL
        print("로그인 완료",EMAIL,PW,R_EMAIL,CHROME_NUMBER)
        
        GMAIL_WB.save(FILE_GMAIL)
        n += 1 #다음 파일 실행 / 엑셀 다음 행렬

print("GMAIL 자동가입 프로그램 종료")