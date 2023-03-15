import os

import pyautogui
import keyboard
import time

import openpyxl as op
import clipboard

os.getcwd()

#==================메타지갑 저장 위한 엑셀 파일 폴더설정==============
PATH_EXCEL = "C:/Users/V/Desktop/다계정/"
FILE_EXCLEL = "PC3_METAMASK_ADDRESS.xlsx"

EXELFILENAME = FILE_EXCLEL

META_WB = op.load_workbook(PATH_EXCEL + FILE_EXCLEL)
META_WS = META_WB.active
#==================메타지갑 저장 위한 엑셀 파일 폴더설정==============

WIDTH, HEIGHT = pyautogui.size()
IDPW = "12121212" #메타지갑 아이디 비번

##크롬 프로필 폴더 위치 역슬래시\ 가 아닌 /슬래시로 폴더 구분
PATH_CHROME = (r"C:/Users/V/Desktop/다계정")

#폴더 내 파일 리스트화
FILE_LIST_RANGE = os.listdir(PATH_CHROME)
FILE_LIST_LNK_RANGE = [file for file in FILE_LIST_RANGE if file.endswith(".lnk")]

CHROME_NUMBER = ""

#파일 실행 함수 정의
def run(path):
    # 디렉토리 설정
    os.chdir(PATH_CHROME)
    # 파일 실행
    os.startfile(path)


#=======================단축키 함수 정의==========================

#time 단축키
def TIME(how):
    pyautogui.sleep(how)


#tab 을 자동으로 눌러주는 함수
def TAB(how):
    for i in range(how) :
        pyautogui.hotkey('tab',interval = 0.02)

#alt + D 를 자동으로 눌러주는 함수
def ALT_D(how):
    for i in range(how) :  
        pyautogui.hotkey('alt', 'd')

#alt + F 를 자동으로 눌러주는 함수
def ALT_F4(how):
    for i in range(how) :
        pyautogui.hotkey('alt', 'f4')

#enter를 자동으로 눌러주는 함수
def ENTER(how):
    for i in range(how) :
        pyautogui.hotkey('enter')

#전체화면 단축키
def FULL_SCREEN(how):
    for i in range(how) :
        pyautogui.hotkey('win' , 'up')

#=======================단축키 함수 부분 끝==========================

#=======================실행 / 반복 부분 ===========================
#몇번부터 실행?
ALL_RANGE = len(FILE_LIST_LNK_RANGE)
n = int(input("몇번부터 시작? 1 ~ %d: " %ALL_RANGE))
print(type(n),n)
n -= 1 #사용자 편의를 위해 0부터 시작이 아닌 1부터 시작

for i in FILE_LIST_LNK_RANGE:
    #특정 파일 링크 - 파일에는 한글 공백 특수기호가 없어야합니다
    if(n < len(FILE_LIST_LNK_RANGE)):
        executeFile = PATH_CHROME + '/' + FILE_LIST_LNK_RANGE[n]
    else:
        break

    #몇번째 크롬계정인지 확인
    CHROME_NUMBER = FILE_LIST_LNK_RANGE[n]
    print(CHROME_NUMBER) #계정 번호 출력

    run(executeFile) #크롬실행
    
    #============================메타마스크 계정 생성==========================

    TIME(1.5) #잠시 대기시간을 주자
    FULL_SCREEN(1)
    ALT_D(1)
    pyautogui.write('https://chrome.google.com/webstore/detail/metamask/nkbihfbeogaeaoehlefnkodbefgpgknn')
    ENTER(1)
    TIME(5)

    TAB(9)
    ENTER(1) #확장프로그램 설치
    TIME(1)

    TAB(1)
    ENTER(1) #확장프로그램추가

    TIME(15) #메타마스크 들어가는데 오래걸린다    

    TAB(1)
    ENTER(1) #METAMASK 시작하기
    TIME(1.5)

    TAB(1)
    ENTER(1) #괜찮습니다
    TIME(2)

    TAB(2)
    ENTER(1) #지갑생성
    TIME(1.5)

    pyautogui.write(IDPW)
    TAB(1)
    pyautogui.write(IDPW)
    TAB(1)
    ENTER(1) #이용약관 동의

    TAB(2)
    ENTER(1) #계정생성
    TIME(6)

    TAB(8)
    ENTER(1) #다음
    TIME(2)

    TAB(2)
    ENTER(1) #다음에 생성
    TIME(2)

    #클립보드 복사 클릭
    pyautogui.click((WIDTH/2), 190, button='left',clicks=1,interval=1)
    time.sleep(.5)

    #==========================엑셀에 클립보드 내용 저장==========================
    META_RESULT = clipboard.paste()

    META_WS.cell(row = (1 + n), column = 1).value = META_RESULT
    META_WS.cell(row = (1 + n), column = 2).value = CHROME_NUMBER #파일 이름 저장
    print ("%s 계정 %s에 저장되었습니다" % (CHROME_NUMBER, FILE_EXCLEL),META_RESULT)
    META_WB.save(EXELFILENAME) #메타지갑 엑셀 저장
    #==========================엑셀에 클립보드 내용 저장==========================

    ALT_F4(1) #크롬종료
    pyautogui.press('down')

    #============================메타마스크 계정 생성==========================


    #한번 더 반복
    n += 1

META_WB.save(EXELFILENAME) #메타지갑 엑셀 저장
print("메타마스크 자동생성 파일 종료")