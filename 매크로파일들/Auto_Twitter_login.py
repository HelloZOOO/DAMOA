import Main as M

import os
import subprocess

import pyautogui
import keyboard
import time

import openpyxl as op
import clipboard

import random
import string

os.getcwd()
WIDTH, HEIGHT = pyautogui.size()

#==================트위터 아이디 비번 엑셀 임포트==============
PATH_EXCEL = "C:/TWITTER_ACCOUNTS/"
FILE_EXCLEL = "PC2_TWITTER.xlsx"

EXELFILENAME = FILE_EXCLEL

TWITTER_WB = op.load_workbook(PATH_EXCEL + FILE_EXCLEL)
TWITTER_WS = TWITTER_WB["IDPW"] #비번 아이디 DB
CHECK_WS = TWITTER_WB["CHECK"] #로그인 O/X 기록 | 로그인여부 | PDPW

MAX_ROW = TWITTER_WS.max_row

print(TWITTER_WB.sheetnames)

#==================트위터 아이디 비번 엑셀 임포트==============

##크롬 프로필 폴더 위치 역슬래시\ 가 아닌 /슬래시로 폴더 구분
PATH_CHROME = (r"C:/CHROME_ACCOUNTS")

#폴더 내 파일 리스트화
FILE_LIST_RANGE = os.listdir(PATH_CHROME)
FILE_LIST_LNK_RANGE = [file for file in FILE_LIST_RANGE if file.endswith(".lnk")]

CHROME_NUMBER = ""

#파일 실행 함수 정의
def RUN(path):
    # 디렉토리 설정
    os.chdir(PATH_CHROME)
    # 파일 실행
    os.startfile(path)

#=======================실행 / 반복 부분 ===========================
#몇번부터 실행?
n = int(input("몇번부터 시작? 1 ~ %d: " %MAX_ROW))
print(type(n),n)
i = n
n -= 1 #사용자 편의를 위해 0부터 시작이 아닌 1부터 시작

for i in range(MAX_ROW):
    if(n < MAX_ROW):
        
        ID = TWITTER_WS.cell(row = (1 + n), column = 1).value #ID 임포트
        PW = TWITTER_WS.cell(row = (1 + n), column = 2).value #PW 임포트
        RE = TWITTER_WS.cell(row = (1 + n), column = 3).value #복구메일 임포트
        RE_PW = TWITTER_WS.cell(row = (1 + n), column = 4).value #복구메일 비번 임포트
        #특정 파일 링크 - 파일에는 한글 공백 특수기호가 없어야합니다
        executeFile = PATH_CHROME + '/' + FILE_LIST_LNK_RANGE[n]
        
        #몇번째 크롬계정인지 확인
        CHROME_NUMBER = FILE_LIST_LNK_RANGE[n]
        print(CHROME_NUMBER) #계정 번호 출력
        
        #00_크롬초기화부분
        M.CLEAR_SCREEN(1)
        RUN(executeFile)
        M.TIME(0.3)
        M.KILL_CHROME()
        M.TIME(0.2)
        RUN(executeFile)
        M.TIME(1.5)
        
        #01_트위터로 이동
        M.NEW_TAB(1)
        M.TIME(0.5)
        M.ALT_D(1)
        M.WRITE('http://twitter.com/home')
        M.ENTER(1)
        M.TIME(4.5)
        
        #02_로그인 X 상태
        M.ALT_D(1)
        M.TIME(0.5)
        M.CTRL_C(1)
        URL_RESULT = clipboard.paste()
        
        #03_01 로그인 되어있을경우
        if 'i/flow/login' in URL_RESULT:
            M.FULL_SCREEN(1) #ON
            M.TIME(1.2)
            M.CLICK_MOSUE(WIDTH/2,540) #로그인 클릭
            M.CTRL_A(1)
            M.WRITE(ID)
            M.ENTER(1)
            M.TIME(2)
            M.CTRL_A(1)
            M.WRITE(PW)
            M.ENTER(1)
            M.TIME(2.5)
            M.CLICK_MOSUE(1588,300) #비밀번호 저장
            M.TIME(2)
            M.CLICK_MOSUE(1200,500) #아니요 이 기기에만 저장
            M.CLICK_MOSUE(WIDTH/2,540) #복구이메일 클릭
            M.CTRL_A(1)
            M.WRITE(RE)
            M.ENTER(1)
            M.TIME(1.6)
        
            #04 로그인 여부 확인부분
            M.FULL_SCREEN(1) #OFF
            M.TIME(0.5)
            M.ALT_D(1)
            M.WRITE('http://twitter.com/home') #다시 트위터 이동
            M.ENTER(1)
            M.TIME(2.5)
            M.ALT_D(1)
            M.CTRL_C(1)
            URL_RESULT = clipboard.paste()
            CHECK_WS.cell(row = (1 + n), column = 2).value = ID
            CHECK_WS.cell(row = (1 + n), column = 3).value = PW
            CHECK_WS.cell(row = (1 + n), column = 4).value = RE
            CHECK_WS.cell(row = (1 + n), column = 5).value = RE_PW
            CHECK_WS.cell(row = (1 + n), column = 6).value = CHROME_NUMBER
            CHECK_WS.cell(row = (1 + n), column = 7).value = URL_RESULT
            
            if '/home' in URL_RESULT:
                CHECK_WS.cell(row = (1 + n), column = 1).value = 1 #로그인 O
                TWITTER_WB.save(EXELFILENAME) #엑셀저장
            else:
                CHECK_WS.cell(row = (1 + n), column = 1).value = 0 #로그인 X
                TWITTER_WB.save(EXELFILENAME) #엑셀저장
        
        #03_02 로그인 O 상태 
        else:
            CHECK_WS.cell(row = (1 + n), column = 1).value = 1 #로그인 O
            CHECK_WS.cell(row = (1 + n), column = 2).value = ID
            CHECK_WS.cell(row = (1 + n), column = 3).value = PW
            CHECK_WS.cell(row = (1 + n), column = 4).value = RE
            CHECK_WS.cell(row = (1 + n), column = 5).value = RE_PW
            CHECK_WS.cell(row = (1 + n), column = 6).value = CHROME_NUMBER
            CHECK_WS.cell(row = (1 + n), column = 7).value = URL_RESULT
            TWITTER_WB.save(EXELFILENAME) #엑셀저장
        
        M.ALT_F4(1)
        #M.KILL_CHROME()
    else:
        break
    
    n += 1 #다음 크롬프로필 파일

TWITTER_WB.save(EXELFILENAME) #엑셀저장
print("트위터 자동 로그인 종료")