import os

import pyautogui
import time

import openpyxl as op
import clipboard

import Main as M

PUBLIC_PATH = "C:/CHROME_ACCOUNTS_2"

#==================비밀 구문이 저장된 엑셀 파일 폴더 설정==============
PW_PATH_EXCEL = PUBLIC_PATH + "/"
PW_FILE_EXCLEL = "PC5_METAMASK_MnemonicCode.xlsx"

PW_EXELFILENAME = PW_FILE_EXCLEL

PW_META_WB = op.load_workbook(PW_PATH_EXCEL + PW_FILE_EXCLEL)
PW_META_WS = PW_META_WB.active
#==================비밀 구문이 저장된 엑셀 파일 폴더 설정==============


#==================메타지갑 저장 위한 엑셀 파일 폴더설정==============
PATH_EXCEL = PUBLIC_PATH + "/"
FILE_EXCLEL = "PC5_METAMASK_ADDRESS.xlsx"

EXELFILENAME = FILE_EXCLEL

META_WB = op.load_workbook(PATH_EXCEL + FILE_EXCLEL)
META_WS = META_WB.active
#==================메타지갑 저장 위한 엑셀 파일 폴더설정==============

WIDTH, HEIGHT = pyautogui.size()
IDPW = "12121212" #메타지갑 아이디 비번

##크롬 프로필 폴더 위치 역슬래시\ 가 아닌 /슬래시로 폴더 구분
#PATH_CHROME = (r"C:/Users/V/Desktop/다계정")
PATH_CHROME = (PUBLIC_PATH)

#폴더 내 파일 리스트화
FILE_LIST_RANGE = os.listdir(PATH_CHROME)
FILE_LIST_LNK_RANGE = [file for file in FILE_LIST_RANGE if file.endswith(".lnk")]

CHROME_NUMBER = ""

#파일 실행 함수 정의
def run(path):
    # 디렉토리 설정
    os.chdir(PATH_CHROME)
    # 파일 실행
    os.startfile(path)

#=======================실행 / 반복 부분 ===========================
#몇번부터 실행?
ALL_RANGE = len(FILE_LIST_LNK_RANGE)
n = int(input("몇번부터 시작? 1 ~ %d: " %ALL_RANGE))
print(type(n),n)
n -= 1 #사용자 편의를 위해 0부터 시작이 아닌 1부터 시작

for i in FILE_LIST_LNK_RANGE:
    #특정 파일 링크 - 파일에는 한글 공백 특수기호가 없어야합니다
    if(n < len(FILE_LIST_LNK_RANGE)):
        executeFile = PATH_CHROME + '/' + FILE_LIST_LNK_RANGE[n]
    else:
        break

    #몇번째 크롬계정인지 확인
    CHROME_NUMBER = FILE_LIST_LNK_RANGE[n]
    print(CHROME_NUMBER) #계정 번호 출력

    run(executeFile) #크롬실행
    
    #============================메타마스크 계정 생성==========================

    M.TIME(1.5) #잠시 대기시간을 주자
    M.ALT_D(1)
    clipboard.copy("https://chrome.google.com/webstore/detail/metamask/nkbihfbeogaeaoehlefnkodbefgpgknn")
    pyautogui.hotkey('ctrl','v')
    M.ENTER(1)
    M.FULL_SCREEN(1)
    M.TIME(5)

    M.TAB(9)
    M.ENTER(1) #확장프로그램 설치
    M.TIME(1)

    M.TAB(1)
    M.ENTER(1) #확장프로그램추가

    M.TIME(10) #메타마스크 들어가는데 오래걸린다    

    M.TAB(8)
    M.ENTER(1) #기존 지갑 가져오기
    #M.TIME(1)

    M.TAB(4)
    M.ENTER(1) #괜찮습니다
    #M.TIME(1)

    #복구구문입력
    MnemonicCode = PW_META_WS.cell(row = (2 + n), column = 3).value
    M.CLIPBOARD(MnemonicCode) #클립보드에 암호구문 복사
    M.TAB(3)
    pyautogui.hotkey('ctrl','v')
    
    #비밀 복구 구문 확인 (원래는 탭 24번 해야함)
    pyautogui.click((WIDTH/2), 950, button='left',clicks=1,interval=1)
    
    M.WRITE(IDPW)
    M.TAB(1)
    M.WRITE(IDPW)
    M.TAB(1)
    M.CLICK_MOSUE(815,625) #비밀번호를 복구할 수 없음을 이해합니다.
    M.TAB(2)
    M.ENTER(1) #내 지갑 가져오기
    M.TIME(1.5)

    M.TAB(3)
    M.ENTER(1) #지갑 생성 성공 확인했습니다!
    M.TIME(0.5)
    
    M.TAB(4)
    M.ENTER(1) #확인
    M.TIME(3)

    #클립보드 복사 클릭
    pyautogui.click((WIDTH/2), 100, button='left',clicks=1,interval=1)
    M.TIME(0.5)

    #==========================엑셀에 클립보드 내용 저장==========================
    META_RESULT = clipboard.paste()

    META_WS.cell(row = (1 + n), column = 1).value = META_RESULT
    META_WS.cell(row = (1 + n), column = 2).value = CHROME_NUMBER #파일 이름 저장
    META_WS.cell(row = (1 + n), column = 3).value = MnemonicCode #암호 구문 저장
    print ("%s 계정 %s에 저장되었습니다" % (CHROME_NUMBER, FILE_EXCLEL),META_RESULT)
    META_WB.save(EXELFILENAME) #메타지갑 엑셀 저장
    #==========================엑셀에 클립보드 내용 저장==========================

    M.ALT_F4(1) #크롬종료

    #============================메타마스크 계정 생성==========================


    #한번 더 반복
    n += 1

META_WB.save(EXELFILENAME) #메타지갑 엑셀 저장
print("메타마스크 자동생성 파일 종료")